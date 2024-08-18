# script KAB version

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re, os

rm_prov = [0, 2, 3]
rm_kab = [0, 1, 3, 4, 5]

def merge_data(folder_path, output_file):
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)  # Hapus sheet default

    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)

            if not is_valid_excel_file(file_path):
                print(f"Skipping invalid or corrupted file: {filename}")
                continue

            print(f"Processing file: {filename}")
            try:
                wb = openpyxl.load_workbook(file_path)
            except Exception as e:
                print(f"Error loading {filename}: {e}")
                continue
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                # Menentukan nama sheet yang sesuai
                base_name = os.path.splitext(filename)[0]
                sheet_title = f"{base_name}_{sheet}"

                # Memastikan nama sheet tidak melebihi 31 karakter
                sheet_title = sheet_title[:31]
                
                # Membuat sheet baru
                new_ws = new_wb.create_sheet(title=sheet_title)
                
                print(f"Processing sheet: {sheet_title}")
                for row in ws.iter_rows(values_only=True):
                    try:
                        new_ws.append(row)
                    except Exception as e:
                        print(f"Error appending row to {new_ws.title}: {e}")
                        continue

    try:
        new_wb.save(output_file)
        print(f"Successfully saved merged file as {output_file}")
    except Exception as e:
        print(f"Error saving file {output_file}: {e}")

def is_valid_excel_file(file_path):
    try:
        with open(file_path, 'rb') as file:
            header = file.read(4)
            return header == b'PK\x03\x04'  # Header untuk file ZIP
    except Exception as e:
        print(f"Error checking file validity: {e}")
        return False

def make_sheet_name(x):
    try:
        c = re.search('jenis_komoditas', x)
        if c:
            sn = re.split(r'_', x)
            v = sn[10]
        else:
            sn = re.split(r'_', x)
            v = sn[9]
        v = v.replace('.xlsx', '')
    except IndexError:
        sn = re.split(r'_', x)
        v = sn[6] + '_' + sn[7]
    return v

def make_xlsx_name(x):
    sn = re.split(r'_', x)
    v = sn[6] + '_' + sn[7]
    v = v.replace('.xlsx', '')
    return v

def check_folders(p):
    if os.path.exists(p): os.makedirs(p)
    else: pass 

folder_data = "D:\\Publikasi UTP\\data prov" # Ganti dengan folder data yang akan diinput
folder_outp = "D:\\Publikasi UTP\\filter prov" # Ganti dengan path folder output hasil run filter

def filter_data(inp, outp, set, kode, nama):

    # Iter nama file XLSX dalam folder
    for index, dirs, files in os.walk(inp):
        for a in files:
            print("Reading %s"%(a))
            getsheet = ''
            ex = pd.ExcelFile(inp + "\\" + a)

            # Generate nama file excel dan nama sheet dari nama file
            for q in ex.sheet_names:
                try:
                    if set == 'prov':
                        c = re.search('kab', q)
                        if c: getsheet = q
                    if set == 'kab':
                        c = re.search('kec', q)
                        if c: getsheet = q
                except:
                    print('Marker sheet tidak ada')

            # Filter data berdasarkan sheet
            df = ex.parse(getsheet)
            try:
                if set == 'prov':
                    df6401 = df.loc[df.iloc[:, 2] == kode]
                elif set == 'kab':
                    df6401 = df.loc[df.iloc[:, 4] == kode]
            except: print("Switcher harus 'prov' untuk tabel provinsi dan 'kab' untuk tabel kabupaten")

            # Generate nilai agregat
            kabsum = df6401.sum()
            df6401 = df6401._append(kabsum, ignore_index=True)
            sz = len(df6401)

            # Menambahkan baris agregat ke tabel dan drop kolom yang tidak diperlukan
            if set == 'prov':
                df6401.iloc[sz-1, 1] = nama
                df6401 = df6401.drop(df6401.columns[rm_prov], axis=1)
            elif set == 'kab':
                df6401.iloc[sz-1, 2] = nama
                df6401 = df6401.drop(df6401.columns[rm_kab], axis=1)

            # Apply format angka dan desimal ke data hasil filter
            for col in range(1, len(df6401.columns)):
                if df6401.iloc[:, col].dtype.kind in 'i':
                    df6401.iloc[:, col] = df6401.iloc[:, col].map('{:,}'.format)
                elif df6401.iloc[:, col].dtype.kind in 'f':
                    df6401.iloc[:, col] = df6401.iloc[:, col].map('{:,.2f}'.format)
            
            df6401.fillna("–", inplace=True)
            df6401.to_excel("%s\\%s_%s.xlsx"%(outp, make_xlsx_name(a), make_sheet_name(a)), sheet_name=make_sheet_name(a))

# contoh pemakaian untuk filter data :
filter_data(inp=folder_data, outp=folder_outp, set='prov', kode=32, nama='JAWA BARAT')
filter_data(inp=folder_data, outp=folder_outp, set='kab', kode=6401, nama='PASER')

# contoh pemakaian untuk reformat file :
reformat("D:\\file_excel_yang_mau_diformat.xlsx")

# contoh pemakaian untuk merging excel menjadi 1 file banyak sheet :
merge_data("folder_yang_akan_dimerge", "nama_file.xlsx")
